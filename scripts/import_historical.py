#!/usr/bin/env python3
"""
Parse historical NBA Playoffs Fantasy Excel scoreboards into per-year JSON.

Reads ~/Downloads/NBA Playoffs <YEAR> Fantasy Scoreboard.xlsx (or --src dir)
and writes data/<year>/{picks,stats,totals}.json.

Excel layout (consistent across 2022–2025):
  Sheet "Scoreboard":     entrant rows; cols vary by year (header detected dynamically).
  Sheet "Individual Scores":
    Row 1: round labels at cols H, P, X, AF (Round 1, Conference Semis, ...)
    Row 2: headers (Seed, Player, Cost, Points/Round, Eliminated?, Total, [Rate], G1-G7, PPG x4)
    Rows 3+: blocks of players grouped by seed, separated by blank rows.
  Sheet "Budget":         player cost lookup (used as fallback for 2022 where name
                          column doesn't embed cost).
"""

import argparse
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


ROUNDS = ["R1", "CSF", "CF", "Finals"]
ROUND_LABELS = {
    "R1": "Round 1",
    "CSF": "Conference Semis",
    "CF": "Conference Finals",
    "Finals": "Finals",
}


def slugify(name: str) -> str:
    s = name.lower().strip()
    s = re.sub(r"[\u2019']", "", s)
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


def split_name_cost(cell) -> tuple[str, float | None]:
    """For 2023+ where name cells are 'Jimmy Butler 21.4'."""
    if cell is None:
        return "", None
    if isinstance(cell, (int, float)):
        return str(cell), None
    s = str(cell).strip()
    m = re.match(r"^(.*?)(?:\s+(\d+(?:\.\d+)?))?$", s)
    if not m:
        return s, None
    name = m.group(1).strip()
    cost = float(m.group(2)) if m.group(2) else None
    return name, cost


def load_budget_lookup(ws) -> dict[str, float]:
    """Build name -> cost map from the Budget sheet (handles 2022 2-col + 2023+ multi-col)."""
    lookup: dict[str, float] = {}
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        # 2022 form: A=name, B=cost
        if isinstance(a, str) and isinstance(b, (int, float)) and not isinstance(c, (int, float)):
            lookup[a.strip().lower()] = float(b)
        # 2023+ form: A="Name X.Y", B=name, C=cost
        elif isinstance(a, str) and isinstance(b, str) and isinstance(c, (int, float)):
            lookup[b.strip().lower()] = float(c)
    return lookup


def parse_individual_scores(ws, budget_lookup: dict[str, float]):
    """
    Returns dict[player_id] -> {
       name, seed, cost, eliminated, total_raw,
       round_ppg: { round: float },
       games: [ { round, game_num, pts } ],
    }
    """
    # Detect column offsets by reading row 2 headers.
    headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    # Find indices of "Total" and "PPG" columns.
    try:
        total_col = headers.index("Total") + 1
    except ValueError:
        total_col = 6  # default

    # PPG columns appear in order for Round 1, Semis, CF, Finals.
    ppg_cols = [i + 1 for i, h in enumerate(headers) if h == "PPG"]
    if len(ppg_cols) != 4:
        raise ValueError(f"Expected 4 PPG columns, got {len(ppg_cols)}: {ppg_cols}")

    # Game columns are the 7 cols immediately preceding each PPG col.
    round_blocks = [(rd, list(range(p - 7, p)), p) for rd, p in zip(ROUNDS, ppg_cols)]

    players: dict[str, dict] = {}
    current_seed: int | None = None

    seed_re = re.compile(r"^(\d+)\s*Seeds?$", re.IGNORECASE)

    for r in range(3, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        if isinstance(a, str):
            m = seed_re.match(a.strip())
            if m:
                current_seed = int(m.group(1))
        if b is None:
            continue
        name, embedded_cost = split_name_cost(b)
        if not name:
            continue
        cost_cell = ws.cell(r, 3).value
        cost = (
            float(cost_cell) if isinstance(cost_cell, (int, float))
            else (embedded_cost if embedded_cost is not None
                  else budget_lookup.get(name.lower()))
        )
        eliminated = (ws.cell(r, 5).value == "X")
        total_raw_cell = ws.cell(r, total_col).value
        total_raw = float(total_raw_cell) if isinstance(total_raw_cell, (int, float)) else 0.0

        round_ppg: dict[str, float] = {}
        games: list[dict] = []
        for rd, game_cols, ppg_col in round_blocks:
            ppg_val = ws.cell(r, ppg_col).value
            if isinstance(ppg_val, (int, float)):
                round_ppg[rd] = float(ppg_val)
            for gi, c in enumerate(game_cols, start=1):
                v = ws.cell(r, c).value
                if isinstance(v, (int, float)):
                    games.append({"round": rd, "game_num": gi, "pts": float(v)})

        pid = slugify(name)
        # Avoid clobbering if a name appears twice (rare edge case): suffix with seed.
        if pid in players and current_seed and players[pid]["seed"] != current_seed:
            pid = f"{pid}-{current_seed}"
        players[pid] = {
            "name": name,
            "seed": current_seed or 0,
            "cost": cost,
            "eliminated": eliminated,
            "total_raw": total_raw,
            "round_ppg": round_ppg,
            "games": games,
        }
    return players


def parse_scoreboard(ws, players_by_id: dict[str, dict]):
    """
    Returns (entrants_picks, totals_rows).
      entrants_picks: list of { name, picks: { "1": {player_id, name, cost}, ... } }
      totals_rows:    list of { name, total, place, remaining, seeds: [...] }  (for totals.json)
    """
    # Header row (row 1) — find index of "Entrant", "Points", "Place", "Remaining"
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    def col_of(label):
        for i, h in enumerate(headers):
            if isinstance(h, str) and h.strip().lower() == label.lower():
                return i + 1
        return None

    entrant_col = col_of("Entrant")
    points_col = col_of("Points")
    place_col = col_of("Place")
    remaining_col = col_of("Remaining")
    if not entrant_col or not points_col:
        raise ValueError("Could not find Entrant/Points columns in Scoreboard")

    # Seed group columns: numeric headers 1..8 in row 1, each followed by Pts/Elim?/Pts+/-
    seed_starts: dict[int, int] = {}
    for i, h in enumerate(headers):
        try:
            sval = int(float(h)) if h is not None else None
        except (ValueError, TypeError):
            sval = None
        if sval and 1 <= sval <= 8:
            seed_starts[sval] = i + 1

    entrants_picks = []
    totals_rows = []

    for r in range(2, ws.max_row + 1):
        ent = ws.cell(r, entrant_col).value
        if not isinstance(ent, str) or not ent.strip():
            continue
        pts_cell = ws.cell(r, points_col).value
        if not isinstance(pts_cell, (int, float)):
            continue
        place = ws.cell(r, place_col).value if place_col else None
        remaining = ws.cell(r, remaining_col).value if remaining_col else None

        picks: dict[str, dict] = {}
        seed_summaries: list[dict] = []
        for seed in range(1, 9):
            if seed not in seed_starts:
                continue
            base = seed_starts[seed]
            name_cell = ws.cell(r, base).value
            pts_seed = ws.cell(r, base + 1).value
            elim_seed = ws.cell(r, base + 2).value
            name, embedded_cost = split_name_cost(name_cell)
            if not name:
                continue
            pid = slugify(name)
            # Match to player record (may have suffix on collision — rare)
            if pid not in players_by_id:
                # Try suffix variants
                for k in players_by_id:
                    if k.startswith(pid + "-"):
                        pid = k
                        break
            cost = embedded_cost
            if pid in players_by_id and players_by_id[pid].get("cost") is not None:
                cost = players_by_id[pid]["cost"]
            picks[str(seed)] = {
                "player_id": pid,
                "name": name,
                "seed": seed,
                "cost": cost,
            }
            seed_summaries.append({
                "seed": seed,
                "player": name,
                "player_id": pid,
                "cost": cost,
                "pts": float(pts_seed) if isinstance(pts_seed, (int, float)) else 0.0,
                "eliminated": elim_seed == "X",
            })

        entrants_picks.append({"name": ent.strip(), "picks": picks})
        totals_rows.append({
            "name": ent.strip(),
            "total": float(pts_cell),
            "place": int(place) if isinstance(place, (int, float)) else None,
            "remaining": int(remaining) if isinstance(remaining, (int, float)) else None,
            "seeds": seed_summaries,
        })

    return entrants_picks, totals_rows


def write_year(year: int, src_dir: Path, out_root: Path):
    src = src_dir / f"NBA Playoffs {year} Fantasy Scoreboard.xlsx"
    if not src.exists():
        print(f"  SKIP: {src} not found")
        return
    print(f"\n=== Importing {year} ===")
    wb = openpyxl.load_workbook(src, data_only=True)
    budget_lookup = load_budget_lookup(wb["Budget"])
    players = parse_individual_scores(wb["Individual Scores"], budget_lookup)
    entrants_picks, totals_rows = parse_scoreboard(wb["Scoreboard"], players)

    # Sort totals by total desc to ease frontend rendering
    totals_rows.sort(key=lambda r: r["total"], reverse=True)

    out_dir = out_root / str(year)
    out_dir.mkdir(parents=True, exist_ok=True)

    picks_doc = {"year": year, "entrants": entrants_picks}
    stats_doc = {
        "year": year,
        "rounds": ROUNDS,
        "round_labels": ROUND_LABELS,
        "players": players,
    }
    totals_doc = {"year": year, "rows": totals_rows}

    (out_dir / "picks.json").write_text(json.dumps(picks_doc, indent=2))
    (out_dir / "stats.json").write_text(json.dumps(stats_doc, indent=2))
    (out_dir / "totals.json").write_text(json.dumps(totals_doc, indent=2))
    print(f"  Wrote {out_dir}/{{picks,stats,totals}}.json — "
          f"{len(entrants_picks)} entrants, {len(players)} players")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", default=str(Path.home() / "Downloads"),
                    help="Directory containing the .xlsx files")
    ap.add_argument("--out", default=None,
                    help="Output data dir (defaults to ../data relative to script)")
    ap.add_argument("--year", type=int, action="append",
                    help="Year(s) to import; default = 2022..2025")
    args = ap.parse_args()

    src_dir = Path(args.src).expanduser()
    out_root = Path(args.out) if args.out else Path(__file__).resolve().parent.parent / "data"
    years = args.year or [2022, 2023, 2024, 2025]
    for y in years:
        write_year(y, src_dir, out_root)


if __name__ == "__main__":
    main()
